#!/usr/bin/env python3
"""Generate a daily / weekly / monthly marketing performance report (.xlsx) for abc.

Usage:
    python generate_report.py --period {daily|weekly|monthly} [--date YYYY-MM-DD] [--root .]

Reads the mkt-* trackers, computes ROAS/CAC/CTR/CPC/CPM/conv-rate and rollups (nothing is stored back),
writes a styled workbook + .md to reports/marketing/<period>/, and prints a short text summary.
"""
import argparse, csv, datetime as dt, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- house style ----
PINK = "EC4899"; WHITE = "FFFFFF"; GREEN = "16A34A"; RED = "DC2626"; AMBER = "D97706"
GREY = "6B7280"; ZEBRA = "FCE7F3"
INR = '"₹"#,##,##0'; PCT = '0.0%'; MULT = '0.00"x"'; NUM = '#,##,##0'
HEAD_FONT = Font(bold=True, color=WHITE, name="Arial", size=11)
TITLE_FONT = Font(bold=True, size=15, name="Arial", color="0F172A")
SUB_FONT = Font(italic=True, color=GREY, name="Arial", size=10)
BOLD = Font(bold=True, name="Arial")
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- targets ----
TARGET_ROAS = 3.0; TARGET_CAC = 1500; TARGET_CTR = 0.015
UNIT_PRICE = 4999


def div(a, b):
    return a / b if b else None


def load(root, rel):
    with open(root / rel, newline="") as f:
        return list(csv.DictReader(f))


def num(r, k):
    v = r.get(k, "") or 0
    try:
        return float(v)
    except ValueError:
        return 0.0


def agg(rows):
    s = sum(num(r, "spend") for r in rows)
    imp = sum(num(r, "impressions") for r in rows)
    clk = sum(num(r, "clicks") for r in rows)
    cv = sum(num(r, "conversions") for r in rows)
    val = sum(num(r, "conversion_value") for r in rows)
    return {
        "spend": s, "impressions": imp, "clicks": clk, "conversions": cv, "value": val,
        "ctr": div(clk, imp), "cpc": div(s, clk), "cpm": div(s * 1000, imp),
        "cac": div(s, cv), "roas": div(val, s), "cvr": div(cv, clk),
    }


def concept_of(ad_set, concepts):
    """Map an ad_set like 'battery-life-a' to (concept, variant) if its stem is a known creative concept."""
    if "-" in ad_set:
        stem, var = ad_set.rsplit("-", 1)
        if stem in concepts:
            return stem, var.upper()
    return None, None


# ---------- styling helpers ----------
def style_title(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title); c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitle); c.font = SUB_FONT
    ws.row_dimensions[1].height = 22


def write_table(ws, top, headers, rows, fmts, roas_idx=None, widths=None):
    """rows: list of cell values. fmts: per-col number_format or None. Returns next free row."""
    for j, h in enumerate(headers, 1):
        c = ws.cell(top, j, h)
        c.font = HEAD_FONT; c.fill = PatternFill("solid", fgColor=PINK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(rows, 1):
        r = top + i
        for j, val in enumerate(row, 1):
            c = ws.cell(r, j, val)
            c.border = BORDER
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=ZEBRA)
            if fmts[j - 1]:
                c.number_format = fmts[j - 1]
                c.alignment = Alignment(horizontal="right")
            if roas_idx is not None and (j - 1) == roas_idx and isinstance(val, (int, float)):
                c.font = Font(bold=True, color=GREEN if val >= TARGET_ROAS else RED, name="Arial")
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(top + 1, 1)
    return top + len(rows) + 2


def fmt_money(v): return f"₹{v:,.0f}" if v is not None else "n/a"
def fmt_roas(v): return f"{v:.2f}x" if v is not None else "n/a"
def fmt_pct(v): return f"{v*100:.1f}%" if v is not None else "n/a"
def fmt_int(v): return f"{int(round(v)):,}" if v is not None else "n/a"
def fmt_tv(v):  # target value: keep fractions, group integers
    if v is None: return "n/a"
    return f"{int(v):,}" if float(v).is_integer() else f"{v:g}"
def fmt_delta(v): return ("▲ " if v >= 0 else "▼ ") + f"{abs(v)*100:.0f}%" if v is not None else "—"


def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "| " + " | ".join("---" for _ in headers) + " |"]
    for r in rows:
        out.append("| " + " | ".join(str(x) for x in r) + " |")
    return "\n".join(out)


# ---------- task engine ----------
# Owners (from the roster): Sandeep = Performance Marketer (paid/budget/bids/data);
# Pooja = Content & Social Lead (creative/influencer); Ops = inventory; Founder = strategy.
SANDEEP, POOJA, OPS, FOUNDER = "Sandeep", "Pooja", "Ops", "Founder"
PRIO_ORDER = {"High": 0, "Med": 1, "Low": 2}


def task(priority, category, owner, text, trigger=None):
    return {"priority": priority, "category": category, "owner": owner, "text": text, "trigger": trigger}


def render_tasks(tasks):
    if not tasks:
        return ["- [ ] Nothing required — metrics within targets; keep the current setup."]
    tasks = sorted(tasks, key=lambda t: PRIO_ORDER.get(t["priority"], 3))
    lines = []
    for t in tasks:
        trg = f" — _trigger: {t['trigger']}_" if t.get("trigger") else ""
        lines.append(f"- [ ] **[{t['priority']} · {t['category']} · {t['owner']}]** {t['text']}{trg}")
    return lines


# ---------- DAILY ----------
def daily(root, anchor, perf, camps):
    day = [r for r in perf if r["date"] == anchor]
    creatives = load(root, "marketing/creatives/creatives-index.csv")
    wb = Workbook(); ws = wb.active; ws.title = "Daily"
    style_title(ws, "abc — Daily Marketing Report", f"{anchor}  ·  pacing & anomalies  ·  target ROAS {TARGET_ROAS:.1f}x", 11)
    headers = ["Campaign", "Platform", "Spend", "Impr", "Clicks", "Conv", "Revenue", "CTR", "CPC", "CAC", "ROAS"]
    fmts = [None, None, INR, NUM, NUM, NUM, INR, PCT, INR, INR, MULT]
    name = {c["campaign_id"]: c["campaign_name"] for c in camps}
    rows = []
    keyset = sorted({(r["campaign_id"], r["platform"]) for r in day})
    for cid, plat in keyset:
        a = agg([r for r in day if r["campaign_id"] == cid and r["platform"] == plat])
        rows.append([name.get(cid, cid), plat, a["spend"], a["impressions"], a["clicks"], a["conversions"],
                     a["value"], a["ctr"], a["cpc"], a["cac"], a["roas"]])
    tot = agg(day)
    rows.append(["TOTAL", "", tot["spend"], tot["impressions"], tot["clicks"], tot["conversions"],
                 tot["value"], tot["ctr"], tot["cpc"], tot["cac"], tot["roas"]])
    nxt = write_table(ws, 4, headers, rows, fmts, roas_idx=10,
                      widths=[26, 10, 12, 10, 9, 7, 12, 9, 9, 10, 9])
    # make TOTAL bold
    for j in range(1, 12):
        ws.cell(nxt - 2, j).font = BOLD

    # pacing table
    ws2 = wb.create_sheet("Pacing & Flags")
    style_title(ws2, "Pacing & Flags", f"month-to-date through {anchor}", 5)
    mtd_start = anchor[:7] + "-01"
    mrows = []
    for c in camps:
        if c["status"] != "ACTIVE":
            continue
        sp = sum(num(r, "spend") for r in perf if r["campaign_id"] == c["campaign_id"] and mtd_start <= r["date"] <= anchor)
        bud = num(c, "total_budget")
        mrows.append([c["campaign_name"], sp, bud, div(sp, bud)])
    write_table(ws2, 4, ["Campaign", "MTD Spend", "Campaign Budget", "Used %"], mrows,
                [None, INR, INR, PCT], widths=[26, 14, 16, 10])

    # flags + action tasks (structured, single source of truth)
    flags, tasks = build_daily_tasks(perf, anchor, camps, day, tot, mrows, creatives)
    flag_msgs = [f["msg"] for f in flags]
    r = 4 + len(mrows) + 2
    ws2.cell(r, 1, "Flags (last 7 days)").font = BOLD; r += 1
    if not flag_msgs:
        ws2.cell(r, 1, "None — all campaigns at or above target."); r += 1
    for m in flag_msgs:
        ws2.cell(r, 1, "• " + m); r += 1
    r += 1
    ws2.cell(r, 1, "Action plan").font = BOLD; r += 1
    for t in sorted(tasks, key=lambda x: PRIO_ORDER.get(x["priority"], 3)):
        ws2.cell(r, 1, f"[{t['priority']}/{t['category']}/{t['owner']}] {t['text']}"); r += 1
    summary = (f"Daily {anchor}: spend {fmt_money(tot['spend'])}, {int(tot['conversions'])} conv, "
               f"revenue {fmt_money(tot['value'])}, ROAS {fmt_roas(tot['roas'])} "
               f"(target {TARGET_ROAS:.1f}x). Flags: {len(flag_msgs)}. Tasks: {len(tasks)}.")

    # ---- exec summary (markdown) ----
    roas_call = "above" if (tot["roas"] or 0) >= TARGET_ROAS else "below"
    ch_rows = []
    for plat in sorted({r["platform"] for r in day}):
        a = agg([r for r in day if r["platform"] == plat])
        ch_rows.append([plat, fmt_money(a["spend"]), fmt_int(a["conversions"]),
                        fmt_money(a["value"]), fmt_roas(a["roas"]), fmt_pct(a["ctr"])])
    md = [f"# abc — Daily Marketing Report · {anchor}",
          "",
          f"_Pacing & anomalies · target ROAS {TARGET_ROAS:.1f}x · generated {dt.date.today().isoformat()}_",
          "",
          "## Headline",
          f"- **Spend:** {fmt_money(tot['spend'])}  ·  **Conversions:** {fmt_int(tot['conversions'])}  ·  **Revenue:** {fmt_money(tot['value'])}",
          f"- **ROAS:** {fmt_roas(tot['roas'])} ({roas_call} the {TARGET_ROAS:.1f}x target)  ·  **CAC:** {fmt_money(tot['cac'])}  ·  **CTR:** {fmt_pct(tot['ctr'])}",
          f"- **Open flags:** {len(flag_msgs)}  ·  **Action items:** {len(tasks)}",
          "",
          "## By channel",
          md_table(["Channel", "Spend", "Conv", "Revenue", "ROAS", "CTR"], ch_rows) if ch_rows else "_No paid activity on this day._",
          "",
          "## Flags"]
    md += ([f"- {m}" for m in flag_msgs] if flag_msgs else ["- None — all campaigns at or above target."])
    md += ["", "## Action plan — today"]
    md += render_tasks(tasks)
    md += ["", f"_Full workbook: `reports/marketing/daily/{anchor}.xlsx`_", ""]
    return wb, summary, "\n".join(md)


def build_flags(perf, anchor, camps):
    """Return structured flags: each dict has msg (display) + fields used to build tasks."""
    a = dt.date.fromisoformat(anchor)
    last7 = [(a - dt.timedelta(days=i)).isoformat() for i in range(7)]
    prev7 = [(a - dt.timedelta(days=i)).isoformat() for i in range(7, 14)]
    flags = []
    targets = {c["campaign_id"]: c for c in camps}
    for cid, c in targets.items():
        win = [r for r in perf if r["campaign_id"] == cid and r["date"] in last7]
        if not win:
            continue
        m = agg(win)
        if m["roas"] is not None and m["spend"] > 1500 and m["roas"] < TARGET_ROAS:
            flags.append({"kind": "roas_low", "campaign": c["campaign_name"], "platform": None,
                          "msg": f"{c['campaign_name']}: 7d ROAS {fmt_roas(m['roas'])} below {TARGET_ROAS:.1f}x target.",
                          "trigger": f"7d ROAS {fmt_roas(m['roas'])} < {TARGET_ROAS:.1f}x"})
        if m["ctr"] is not None and m["ctr"] < TARGET_CTR:
            flags.append({"kind": "ctr_low", "campaign": c["campaign_name"], "platform": None,
                          "msg": f"{c['campaign_name']}: 7d CTR {fmt_pct(m['ctr'])} below {TARGET_CTR*100:.1f}%.",
                          "trigger": f"7d CTR {fmt_pct(m['ctr'])} < {TARGET_CTR*100:.1f}%"})
    for cid, c in targets.items():
        for plat in {r["platform"] for r in perf if r["campaign_id"] == cid}:
            today = agg([r for r in perf if r["campaign_id"] == cid and r["platform"] == plat and r["date"] == anchor])
            base = agg([r for r in perf if r["campaign_id"] == cid and r["platform"] == plat and r["date"] in prev7])
            if today["cpm"] and base["cpm"] and today["cpm"] > 1.5 * base["cpm"]:
                flags.append({"kind": "cpm_spike", "campaign": c["campaign_name"], "platform": plat,
                              "msg": f"{c['campaign_name']} / {plat}: CPM spike {fmt_money(today['cpm'])} vs 7d avg {fmt_money(base['cpm'])}.",
                              "trigger": f"CPM {fmt_money(today['cpm'])} > 1.5x 7d avg {fmt_money(base['cpm'])}"})
            if today["spend"] > 300 and today["conversions"] == 0:
                flags.append({"kind": "no_conv", "campaign": c["campaign_name"], "platform": plat,
                              "msg": f"{c['campaign_name']} / {plat}: {fmt_money(today['spend'])} spent, 0 conversions on {anchor}.",
                              "trigger": f"{fmt_money(today['spend'])} spend, 0 conversions"})
    return flags


def build_daily_tasks(perf, anchor, camps, day, tot, mrows, creatives):
    """Action list for the day: flag-driven fixes + pacing + data/creative/inventory nudges."""
    tasks = []
    flags = build_flags(perf, anchor, camps)
    for f in flags:
        scope = f["campaign"] + (f" / {f['platform']}" if f["platform"] else "")
        if f["kind"] == "roas_low":
            tasks.append(task("High", "Budget", SANDEEP, f"Trim or pause underperformers in {scope}; reallocate to top-ROAS campaign.", f["trigger"]))
        elif f["kind"] == "no_conv":
            tasks.append(task("High", "Budget", SANDEEP, f"Pause {scope} or verify conversion tracking — spend with no conversions.", f["trigger"]))
        elif f["kind"] == "cpm_spike":
            tasks.append(task("Med", "Creative", POOJA, f"Rotate creative / widen audience on {scope} (frequency fatigue).", f["trigger"]))
        elif f["kind"] == "ctr_low":
            tasks.append(task("Med", "Creative", POOJA, f"Refresh the lead creative for {scope} — CTR under target.", f["trigger"]))
    # pacing vs monthly budget
    a = dt.date.fromisoformat(anchor)
    days_in_month = (dt.date(a.year + (a.month == 12), (a.month % 12) + 1, 1) - dt.timedelta(days=1)).day
    mtd_spend = sum(r[1] for r in mrows); mtd_budget = sum(r[2] for r in mrows)
    if mtd_budget:
        projected = mtd_spend / a.day * days_in_month
        if projected > mtd_budget * 1.05:
            tasks.append(task("High", "Budget", SANDEEP, f"Lower daily caps — MTD pace projects {fmt_money(projected)} vs {fmt_money(mtd_budget)} budget.",
                              f"projected {fmt_money(projected)} > budget {fmt_money(mtd_budget)}"))
        elif projected < mtd_budget * 0.8:
            tasks.append(task("Low", "Budget", SANDEEP, f"Underpacing ({fmt_money(projected)} vs {fmt_money(mtd_budget)}) — room to scale winning ad-sets.",
                              f"projected {fmt_money(projected)} < 80% of budget"))
    # best channel today
    ch = [(plat, agg([r for r in day if r["platform"] == plat])) for plat in {r["platform"] for r in day}]
    ch = [(p, m) for p, m in ch if m["roas"] is not None]
    if len(ch) >= 2:
        best = max(ch, key=lambda x: x[1]["roas"])
        tasks.append(task("Med", "Budget", SANDEEP, f"Shift incremental budget to {best[0]} — today's best ROAS ({fmt_roas(best[1]['roas'])}).",
                          f"{best[0]} ROAS {fmt_roas(best[1]['roas'])}"))
    # DRAFT creatives waiting
    drafts = [c for c in creatives if c.get("status") == "DRAFT"]
    if drafts:
        tasks.append(task("Med", "Creative", POOJA, f"Launch or discard {len(drafts)} DRAFT creative(s) in the library.",
                          f"{len(drafts)} creative(s) status=DRAFT"))
    # data completeness: channel active in last 7d but no row today
    last7 = {(a - dt.timedelta(days=i)).isoformat() for i in range(1, 8)}
    ran_recently = {r["platform"] for r in perf if r["date"] in last7}
    ran_today = {r["platform"] for r in day}
    missing = ran_recently - ran_today
    if missing:
        tasks.append(task("Med", "Data", SANDEEP, f"Confirm {anchor} numbers imported for: {', '.join(sorted(missing))}.",
                          f"no rows today for {', '.join(sorted(missing))}"))
    # inventory nudge on a sales-pace spike
    trailing = [r for r in perf if r["date"] in last7]
    avg_conv = (sum(num(r, "conversions") for r in trailing) / 7) if trailing else 0
    if avg_conv and tot["conversions"] > 1.3 * avg_conv:
        tasks.append(task("Low", "Inventory", OPS, "Check finished-goods stock for AQM-PRO-01 — sales pace running above the 7-day average.",
                          f"{fmt_int(tot['conversions'])} conv vs 7d avg {avg_conv:.1f}"))
    return flags, tasks


# ---------- WEEKLY ----------
def weekly(root, anchor, perf, camps, infl, creatives):
    a = dt.date.fromisoformat(anchor)
    iso = a.isocalendar()
    mon = a - dt.timedelta(days=a.weekday())
    sun = mon + dt.timedelta(days=6)
    pmon = mon - dt.timedelta(days=7); psun = mon - dt.timedelta(days=1)
    label = f"{iso[0]}-W{iso[1]:02d}"
    def inrange(r, s, e): return s.isoformat() <= r["date"] <= e.isoformat()
    cur = [r for r in perf if inrange(r, mon, sun)]
    prev = [r for r in perf if inrange(r, pmon, psun)]

    wb = Workbook(); ws = wb.active; ws.title = "By Channel"
    style_title(ws, f"abc — Weekly Marketing Review ({label})", f"{mon} to {sun}  ·  WoW vs {pmon}–{psun}", 8)
    headers = ["Channel", "Spend", "Revenue", "Conv", "ROAS", "CAC", "CTR", "Spend WoW"]
    fmts = [None, INR, INR, NUM, MULT, INR, PCT, PCT]
    rows = []
    for plat in sorted({r["platform"] for r in cur}):
        m = agg([r for r in cur if r["platform"] == plat])
        pm = agg([r for r in prev if r["platform"] == plat])
        wow = div(m["spend"] - pm["spend"], pm["spend"]) if pm["spend"] else None
        rows.append([plat, m["spend"], m["value"], m["conversions"], m["roas"], m["cac"], m["ctr"], wow])
    tot = agg(cur); ptot = agg(prev)
    rows.append(["BLENDED", tot["spend"], tot["value"], tot["conversions"], tot["roas"], tot["cac"], tot["ctr"],
                 div(tot["spend"] - ptot["spend"], ptot["spend"]) if ptot["spend"] else None])
    nxt = write_table(ws, 4, headers, rows, fmts, roas_idx=4, widths=[12, 12, 12, 8, 9, 10, 9, 11])
    for j in range(1, 9):
        ws.cell(nxt - 2, j).font = BOLD

    # campaigns vs target
    wsc = wb.create_sheet("Campaigns")
    style_title(wsc, "Campaigns vs target", label, 6)
    name = {c["campaign_id"]: c for c in camps}
    crows = []
    for cid in sorted({r["campaign_id"] for r in cur}):
        m = agg([r for r in cur if r["campaign_id"] == cid])
        c = name.get(cid, {})
        crows.append([c.get("campaign_name", cid), m["spend"], m["value"], m["roas"],
                      c.get("target_metric", ""), num(c, "target_value")])
    write_table(wsc, 4, ["Campaign", "Spend", "Revenue", "ROAS", "Target", "Target val"], crows,
                [None, INR, INR, MULT, None, NUM], roas_idx=3, widths=[26, 12, 12, 9, 12, 11])

    # creative leaderboard
    wsx = wb.create_sheet("Creatives")
    style_title(wsx, "Creative leaderboard (by ROAS)", f"{label}  ·  Meta creatives", 6)
    concepts = {c["concept"] for c in creatives}
    bucket = {}
    for r in cur:
        cpt, var = concept_of(r["ad_set"], concepts)
        if not cpt:
            continue
        k = (cpt, var)
        b = bucket.setdefault(k, {"spend": 0, "value": 0, "conv": 0})
        b["spend"] += num(r, "spend"); b["value"] += num(r, "conversion_value"); b["conv"] += num(r, "conversions")
    lead = []
    for (cpt, var), b in bucket.items():
        lead.append([cpt, var, b["spend"], b["value"], b["conv"], div(b["value"], b["spend"])])
    lead.sort(key=lambda x: (x[5] is None, -(x[5] or 0)))
    write_table(wsx, 4, ["Concept", "Variant", "Spend", "Revenue", "Conv", "ROAS"], lead,
                [None, None, INR, INR, NUM, MULT], roas_idx=5, widths=[18, 9, 12, 12, 8, 9])

    # influencer this week
    wsi = wb.create_sheet("Influencer")
    style_title(wsi, "Influencer collabs live this week", label, 6)
    irows = []
    for c in infl:
        gl = c.get("go_live_date", "")
        if gl and mon.isoformat() <= gl <= sun.isoformat():
            oa = num(c, "orders_attributed")
            irows.append([c["influencer_name"], c["handle"], c["deliverable"], num(c, "fee"),
                          oa, div(num(c, "fee"), oa)])
    if not irows:
        irows = [["— none live this week —", "", "", 0, 0, None]]
    write_table(wsi, 4, ["Influencer", "Handle", "Deliverable", "Fee", "Orders", "Cost/Order"], irows,
                [None, None, None, INR, NUM, INR], widths=[20, 20, 12, 11, 9, 12])

    top = lead[0] if lead else None
    summary = (f"Week {label}: spend {fmt_money(tot['spend'])}, revenue {fmt_money(tot['value'])}, "
               f"blended ROAS {fmt_roas(tot['roas'])}, CAC {fmt_money(tot['cac']) if tot['cac'] else 'n/a'}. "
               + (f"Top creative: {top[0]}-{top[1]} at {fmt_roas(top[5])}." if top else ""))

    # ---- exec summary (markdown) ----
    spend_wow = div(tot["spend"] - ptot["spend"], ptot["spend"]) if ptot["spend"] else None
    rev_wow = div(tot["value"] - ptot["value"], ptot["value"]) if ptot["value"] else None
    # best/worst campaign by ROAS this week
    camp_perf = []
    for cid in {r["campaign_id"] for r in cur}:
        a = agg([r for r in cur if r["campaign_id"] == cid])
        camp_perf.append((name.get(cid, {}).get("campaign_name", cid), a["roas"]))
    camp_perf = [c for c in camp_perf if c[1] is not None]
    camp_perf.sort(key=lambda x: x[1], reverse=True)
    best = camp_perf[0] if camp_perf else None
    worst = camp_perf[-1] if camp_perf else None
    ch_rows = []
    for plat in sorted({r["platform"] for r in cur}):
        a = agg([r for r in cur if r["platform"] == plat])
        ch_rows.append([plat, fmt_money(a["spend"]), fmt_money(a["value"]), fmt_roas(a["roas"]),
                        fmt_money(a["cac"]), fmt_pct(a["ctr"])])
    lead_rows = [[c, v, fmt_money(s), fmt_money(val), fmt_int(cv), fmt_roas(ro)]
                 for c, v, s, val, cv, ro in lead[:5]]
    live_infl = [c for c in infl if c.get("go_live_date", "") and mon.isoformat() <= c["go_live_date"] <= sun.isoformat()]

    md = [f"# abc — Weekly Marketing Review · {label}",
          "",
          f"_{mon} to {sun} · WoW vs {pmon}–{psun} · generated {dt.date.today().isoformat()}_",
          "",
          "## Headline",
          f"- **Spend:** {fmt_money(tot['spend'])} ({fmt_delta(spend_wow)} WoW)  ·  **Revenue:** {fmt_money(tot['value'])} ({fmt_delta(rev_wow)} WoW)",
          f"- **Blended ROAS:** {fmt_roas(tot['roas'])} (target {TARGET_ROAS:.1f}x)  ·  **CAC:** {fmt_money(tot['cac'])}  ·  **Conversions:** {fmt_int(tot['conversions'])}",
          (f"- **Best campaign:** {best[0]} ({fmt_roas(best[1])})  ·  **Needs attention:** {worst[0]} ({fmt_roas(worst[1])})" if best else ""),
          "",
          "## By channel",
          md_table(["Channel", "Spend", "Revenue", "ROAS", "CAC", "CTR"], ch_rows) if ch_rows else "_No paid activity this week._",
          "",
          "## Creative leaderboard (Meta, by ROAS)",
          md_table(["Concept", "Variant", "Spend", "Revenue", "Conv", "ROAS"], lead_rows) if lead_rows else "_No creative-tagged spend this week._",
          "",
          "## Influencer live this week"]
    if live_infl:
        for c in live_infl:
            oa = num(c, "orders_attributed")
            md.append(f"- **{c['influencer_name']}** ({c['handle']}, {c['deliverable']}) — fee {fmt_money(num(c,'fee'))}, "
                      f"{fmt_int(oa)} orders, cost/order {fmt_money(div(num(c,'fee'), oa))}")
    else:
        md.append("- None went live this week.")
    # ---- weekly action plan ----
    wtasks = []
    if tot["roas"] is not None and tot["roas"] < TARGET_ROAS:
        wtasks.append(task("High", "Budget", SANDEEP, "Blended ROAS under target — cut the weakest ad-sets and consolidate spend on winners.",
                           f"blended ROAS {fmt_roas(tot['roas'])} < {TARGET_ROAS:.1f}x"))
    if best and worst and best[0] != worst[0]:
        wtasks.append(task("High", "Budget", SANDEEP, f"Reallocate budget from {worst[0]} to {best[0]} for next week.",
                           f"{best[0]} {fmt_roas(best[1])} vs {worst[0]} {fmt_roas(worst[1])}"))
    if lead:
        t = lead[0]
        wtasks.append(task("Med", "Creative", POOJA, f"Scale the {t[0]}-{t[1]} ad-set and brief 2 new variants of this concept.",
                           f"top creative ROAS {fmt_roas(t[5])}"))
        dead = [l for l in lead if (l[5] or 0) == 0 and l[2] > 0]
        for d in dead:
            wtasks.append(task("Med", "Creative", POOJA, f"Pause/retire {d[0]}-{d[1]} — spend with no return.",
                               f"{fmt_money(d[2])} spend, 0 revenue"))
    for plat in sorted({r["platform"] for r in cur}):
        a = agg([r for r in cur if r["platform"] == plat])
        if a["cac"] is not None and a["cac"] > TARGET_CAC:
            wtasks.append(task("Med", "Budget", SANDEEP, f"Tighten bids / targeting on {plat} — CAC above target.",
                               f"{plat} CAC {fmt_money(a['cac'])} > {fmt_money(TARGET_CAC)}"))
    for c in live_infl:
        if c.get("status") == "COMPLETED" and num(c, "orders_attributed") == 0:
            wtasks.append(task("Med", "Influencer", POOJA, f"Fill orders_attributed for {c['influencer_name']} from coupon {c.get('promo_code') or 'n/a'}.",
                               "completed collab, orders not recorded"))
        if c.get("status") in ("BOOKED", "NEGOTIATING"):
            wtasks.append(task("Low", "Influencer", POOJA, f"Confirm go-live & deliverables with {c['influencer_name']}.", f"status {c['status']}"))
    wtasks.append(task("Low", "Data", SANDEEP, "Run organize_creatives.py to index any new assets added this week.", "weekly housekeeping"))

    md += ["", "## Action plan — this week"]
    md += render_tasks(wtasks)
    md += ["", f"_Full workbook: `reports/marketing/weekly/{label}.xlsx`_", ""]
    return wb, summary, "\n".join([m for m in md if m is not None])


# ---------- MONTHLY ----------
def monthly(root, anchor, perf, camps, infl):
    ym = anchor[:7]
    y, m = int(ym[:4]), int(ym[5:7])
    pm = f"{y-1}-12" if m == 1 else f"{y}-{m-1:02d}"
    cur = [r for r in perf if r["date"][:7] == ym]
    prev = [r for r in perf if r["date"][:7] == pm]
    tot = agg(cur); ptot = agg(prev)

    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    style_title(ws, f"abc — Monthly Marketing Report ({ym})", f"the board number  ·  vs {pm}", 4)
    def delta(c, p): return div(c - p, p) if p else None
    srows = [
        ["Total spend", tot["spend"], ptot["spend"], delta(tot["spend"], ptot["spend"])],
        ["Attributed revenue", tot["value"], ptot["value"], delta(tot["value"], ptot["value"])],
        ["Conversions", tot["conversions"], ptot["conversions"], delta(tot["conversions"], ptot["conversions"])],
        ["Blended ROAS", tot["roas"], ptot["roas"], delta(tot["roas"] or 0, ptot["roas"] or 0)],
        ["Blended CAC", tot["cac"], ptot["cac"], delta(tot["cac"] or 0, ptot["cac"] or 0)],
    ]
    fmts = [None, INR, INR, PCT]
    # ROAS row uses MULT for value/prev cols — handle by per-row format
    nxt = write_table(ws, 4, ["Metric", ym, pm, "Δ MoM"], srows, fmts, widths=[22, 14, 14, 10])
    ws.cell(4 + 4, 2).number_format = MULT; ws.cell(4 + 4, 3).number_format = MULT  # ROAS row

    # by channel
    wsc = wb.create_sheet("By Channel")
    style_title(wsc, "Spend & ROAS by channel", ym, 6)
    crows = []
    for plat in sorted({r["platform"] for r in cur}):
        a = agg([r for r in cur if r["platform"] == plat])
        crows.append([plat, a["spend"], div(a["spend"], tot["spend"]), a["value"], a["conversions"], a["roas"]])
    write_table(wsc, 4, ["Channel", "Spend", "Spend mix", "Revenue", "Conv", "ROAS"], crows,
                [None, INR, PCT, INR, NUM, MULT], roas_idx=5, widths=[12, 13, 10, 13, 8, 9])

    # campaign scorecard
    wss = wb.create_sheet("Campaign Scorecard")
    style_title(wss, "Campaign scorecard — target vs actual", ym, 7)
    name = {c["campaign_id"]: c for c in camps}
    rows = []
    for cid in sorted({r["campaign_id"] for r in cur}):
        a = agg([r for r in cur if r["campaign_id"] == cid])
        c = name.get(cid, {})
        tgt = c.get("target_metric", ""); tval = num(c, "target_value")
        actual = {"ROAS": a["roas"], "CAC": a["cac"], "REVENUE": a["value"],
                  "CONVERSIONS": a["conversions"],
                  "CTR": (a["ctr"] * 100 if a["ctr"] is not None else None)}.get(tgt)
        hit = "—"
        if actual is not None and tval:
            if tgt == "CAC":
                hit = "HIT" if actual <= tval else "MISS"
            else:
                hit = "HIT" if actual >= tval else "MISS"
        rows.append([c.get("campaign_name", cid), a["spend"], a["value"], a["roas"], tgt, tval, hit])
    nxt = write_table(wss, 4, ["Campaign", "Spend", "Revenue", "ROAS", "Target", "Target val", "Result"], rows,
                      [None, INR, INR, MULT, None, NUM, None], roas_idx=3, widths=[26, 12, 12, 9, 12, 11, 9])
    for i in range(len(rows)):
        cell = wss.cell(5 + i, 7)
        cell.font = Font(bold=True, color=GREEN if cell.value == "HIT" else RED if cell.value == "MISS" else GREY, name="Arial")
        cell.alignment = Alignment(horizontal="center")

    # influencer month
    wsi = wb.create_sheet("Influencer")
    style_title(wsi, "Influencer collabs this month", ym, 6)
    irows = []
    for c in infl:
        gl = c.get("go_live_date", "")
        if gl[:7] == ym:
            oa = num(c, "orders_attributed")
            irows.append([c["influencer_name"], c["deliverable"], c["status"], num(c, "fee"), oa, div(num(c, "fee"), oa)])
    if not irows:
        irows = [["— none this month —", "", "", 0, 0, None]]
    write_table(wsi, 4, ["Influencer", "Deliverable", "Status", "Fee", "Orders", "Cost/Order"], irows,
                [None, None, None, INR, NUM, INR], widths=[20, 12, 12, 11, 9, 12])

    summary = (f"Month {ym}: spend {fmt_money(tot['spend'])}, revenue {fmt_money(tot['value'])}, "
               f"blended ROAS {fmt_roas(tot['roas'])}, CAC {fmt_money(tot['cac']) if tot['cac'] else 'n/a'}, "
               f"{int(tot['conversions'])} conversions.")

    # ---- exec summary (markdown) ----
    spend_mom = delta(tot["spend"], ptot["spend"])
    rev_mom = delta(tot["value"], ptot["value"])
    roas_mom = delta(tot["roas"] or 0, ptot["roas"] or 0)
    ch_rows = [[plat, fmt_money(sp), fmt_pct(mix), fmt_money(rv), fmt_int(cv), fmt_roas(ro)]
               for plat, sp, mix, rv, cv, ro in crows]
    score_rows = [[nm, fmt_money(sp), fmt_money(rv), fmt_roas(ro), f"{tgt} {fmt_tv(tv)}", res]
                  for nm, sp, rv, ro, tgt, tv, res in rows]
    hits = sum(1 for r in rows if r[6] == "HIT"); misses = sum(1 for r in rows if r[6] == "MISS")
    month_infl = [c for c in infl if c.get("go_live_date", "")[:7] == ym]
    infl_orders = sum(num(c, "orders_attributed") for c in month_infl)
    infl_fee = sum(num(c, "fee") for c in month_infl)
    md = [f"# abc — Monthly Marketing Report · {ym}",
          "",
          f"_The board number · vs {pm} · generated {dt.date.today().isoformat()}_",
          "",
          "## Headline",
          f"- **Spend:** {fmt_money(tot['spend'])} ({fmt_delta(spend_mom)} MoM)  ·  **Revenue:** {fmt_money(tot['value'])} ({fmt_delta(rev_mom)} MoM)",
          f"- **Blended ROAS:** {fmt_roas(tot['roas'])} ({fmt_delta(roas_mom)} MoM, target {TARGET_ROAS:.1f}x)  ·  **CAC:** {fmt_money(tot['cac'])}  ·  **Conversions:** {fmt_int(tot['conversions'])}",
          f"- **Campaign targets:** {hits} hit / {misses} missed",
          "",
          "## Spend & ROAS by channel",
          md_table(["Channel", "Spend", "Mix", "Revenue", "Conv", "ROAS"], ch_rows) if ch_rows else "_No paid activity._",
          "",
          "## Campaign scorecard",
          md_table(["Campaign", "Spend", "Revenue", "ROAS", "Target", "Result"], score_rows) if score_rows else "_No campaigns ran._",
          "",
          "## Influencer",
          (f"- {len(month_infl)} collab(s) this month — total fee {fmt_money(infl_fee)}, {fmt_int(infl_orders)} attributed orders."
           if month_infl else "- No influencer activity this month."),
          "",
          "## Takeaways",
          f"- Blended ROAS of {fmt_roas(tot['roas'])} is {'comfortably above' if (tot['roas'] or 0) >= TARGET_ROAS else 'below'} the {TARGET_ROAS:.1f}x target.",
          (f"- Strongest channel: {max(crows, key=lambda r: r[5] or 0)[0]}; "
           f"largest spend: {max(crows, key=lambda r: r[1])[0]}." if crows else ""),
          (f"- {misses} campaign(s) missed target — review before scaling." if misses else "- All campaigns met target."),
          ""]

    # ---- monthly action plan ----
    nm_y, nm_m = (y + 1, 1) if m == 12 else (y, m + 1)
    nm = f"{nm_y}-{nm_m:02d}"
    mtasks = []
    if crows:
        top_ch = max(crows, key=lambda r: r[5] or 0)[0]
        low_ch = min(crows, key=lambda r: r[5] if r[5] is not None else 1e9)[0]
        mtasks.append(task("High", "Budget", SANDEEP, f"Set the {nm} channel budget — weight toward {top_ch}, trim {low_ch}.",
                           f"{top_ch} best ROAS this month"))
    for r in rows:
        if r[6] == "MISS":
            mtasks.append(task("High", "Budget", SANDEEP, f"Restructure or pause {r[0]} — missed its {r[4]} target.",
                               f"{r[0]} missed {r[4]} {fmt_tv(r[5])}"))
    mtasks.append(task("Med", "Creative", POOJA, "Build next month's creative slate from this month's winning concepts; retire stale assets.",
                       "monthly creative refresh"))
    mtasks.append(task("Med", "Influencer", POOJA, "Book the next batch of collabs and fill orders_attributed for completed ones.",
                       "influencer pipeline"))
    for c in camps:
        if c.get("status") == "PLANNED" and c.get("start_date"):
            mtasks.append(task("Low", "Strategy", FOUNDER, f"Prep {c['campaign_name']} (creative + influencer) ahead of {c['start_date']}.",
                               f"planned campaign starts {c['start_date']}"))
    if tot["roas"] is not None and tot["roas"] > 1.5 * TARGET_ROAS:
        mtasks.append(task("Low", "Strategy", FOUNDER, f"Consider raising the ROAS target or increasing budget — running well above {TARGET_ROAS:.1f}x.",
                           f"blended ROAS {fmt_roas(tot['roas'])}"))
    mtasks.append(task("Low", "Strategy", FOUNDER, f"Validate the seasonality assumption against {ym} sales data.", "seasonality TBD in brand profile"))

    md += ["## Action plan — this month"]
    md += render_tasks(mtasks)
    md += ["", f"_Full workbook: `reports/marketing/monthly/{ym}.xlsx`_", ""]
    return wb, summary, "\n".join([m for m in md if m is not None])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--period", required=True, choices=["daily", "weekly", "monthly"])
    ap.add_argument("--date")
    ap.add_argument("--root", default=".")
    args = ap.parse_args()
    root = Path(args.root).resolve()

    perf = load(root, "trackers/mkt-ad-performance/sample-data.csv")
    camps = load(root, "trackers/mkt-campaigns/sample-data.csv")
    infl = load(root, "trackers/mkt-influencer-collabs/sample-data.csv")
    creatives = load(root, "marketing/creatives/creatives-index.csv")
    if not perf:
        print("No ad-performance rows found."); return 1
    anchor = args.date or max(r["date"] for r in perf)

    if args.period == "daily":
        wb, summary, md = daily(root, anchor, perf, camps)
        out = root / f"reports/marketing/daily/{anchor}.xlsx"
    elif args.period == "weekly":
        wb, summary, md = weekly(root, anchor, perf, camps, infl, creatives)
        iso = dt.date.fromisoformat(anchor).isocalendar()
        out = root / f"reports/marketing/weekly/{iso[0]}-W{iso[1]:02d}.xlsx"
    else:
        wb, summary, md = monthly(root, anchor, perf, camps, infl)
        out = root / f"reports/marketing/monthly/{anchor[:7]}.xlsx"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    md_out = out.with_suffix(".md")
    md_out.write_text(md, encoding="utf-8")
    print(summary)
    print(f"WROTE {out}")
    print(f"WROTE {md_out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
